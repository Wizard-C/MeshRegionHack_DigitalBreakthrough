'''
XLS
'''
import xml.etree.ElementTree as ET
import datetime
from openpyxl import Workbook
import os

directory = os.path.abspath(os.curdir)  # Директория
# Создаем подключение
wb = Workbook()
ws = wb.active

# Назначаем данные нужной ячейке
ws['A1'] = '42'
# ws['B2'] = 24
# ws.append([1, 2, 3])
# ws['A2'] = datetime.datetime.now()

# Save the file
wb.save("sample.xml")


'''
XML
'''

# This is the parent (root) tag
# onto which other tags would be
# created
data = ET.Element('chess')

# Adding a subtag named `Opening`
# inside our root tag
element1 = ET.SubElement(data, 'Opening')

# Adding subtags under the `Opening`
# subtag
s_elem1 = ET.SubElement(element1, 'E4')
s_elem2 = ET.SubElement(element1, 'D4')

# Adding attributes to the tags under
# `items`
s_elem1.set('type', 'Accepted')
s_elem2.set('type', 'Declined')

# Adding text between the `E4` and `D5`
# subtag
s_elem1.text = "King's Gambit Accepted"
s_elem2.text = "Queen's Gambit Declined"

# Converting the xml data to byte object,
# for allowing flushing data to file
# stream
b_xml = ET.tostring(data)

# Opening a file under the name `items2.xml`,
# with operation mode `wb` (write + binary)
with open("sample.xml", "wb") as f:
    f.write(b_xml)
